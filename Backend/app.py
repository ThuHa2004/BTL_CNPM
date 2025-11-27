from flask import Flask, request, jsonify, render_template, session, redirect, url_for, send_file
from db import get_connection
from datetime import datetime
import os
from io import BytesIO

# Cấu hình thư viện xuất Excel (Bắt buộc cài đặt: pip install openpyxl)
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

app = Flask(__name__, template_folder='templates', static_folder='static')
# LƯU Ý: Đổi chuỗi này thành mã bí mật ngẫu nhiên để bảo mật session
app.secret_key = 'your-super-secret-key-change-this'


# ==========================================
# 1. AUTHENTICATION & NAVIGATION ROUTES
# ==========================================

@app.route('/')
@app.route('/login')
def login_page():
    if 'user' in session:
        user = session['user']
        if user.get('user_type') == 'teacher':
            return redirect(url_for('teacher_dashboard'))
        else:
            return redirect(url_for('dashboard'))
    return render_template('login.html')


@app.route('/login', methods=['POST'])
def login():
    data = request.get_json()
    ten_dang_nhap = data.get('ten_dang_nhap')
    mat_khau = data.get('mat_khau')

    if not ten_dang_nhap or not mat_khau:
        return jsonify({'status': 'error', 'message': 'Vui lòng điền đầy đủ thông tin'}), 400

    conn = None
    try:
        conn = get_connection()
        if not conn:
            return jsonify({'status': 'error', 'message': 'Lỗi kết nối Database Server'}), 500
        cursor = conn.cursor()

        # 1. Kiểm tra Sinh Viên
        cursor.execute("""
            SELECT sv.MaSinhVien, sv.HoTen, sv.Email, sv.SoDienThoai, sv.GioiTinh, sv.NgaySinh, sv.MaLop
            FROM TaiKhoanSinhVien tks
            JOIN SinhVien sv ON tks.MaSinhVien = sv.MaSinhVien
            WHERE tks.TenDangNhap = ? AND tks.MatKhau = ?
        """, (ten_dang_nhap, mat_khau))
        sinh_vien = cursor.fetchone()

        if sinh_vien:
            user_data = {
                'user_type': 'student',
                'ma_sinh_vien': sinh_vien[0],
                'ho_ten': sinh_vien[1],
                'email': sinh_vien[2],
                'so_dien_thoai': sinh_vien[3],
                'gioi_tinh': sinh_vien[4],
                'ngay_sinh': sinh_vien[5].strftime('%Y-%m-%d') if sinh_vien[5] else '',
                'ma_lop': sinh_vien[6]
            }
            session['user'] = user_data
            return jsonify(
                {'status': 'success', 'message': 'Đăng nhập thành công', 'user': user_data, 'redirect': '/dashboard'})

        # 2. Kiểm tra Giảng Viên
        cursor.execute("""
            SELECT gv.MaGiaoVien, gv.HoTen, gv.Email, gv.SoDienThoai, gv.GioiTinh, gv.NgaySinh, gv.ChucVu
            FROM TaiKhoanGiaoVien tkg
            JOIN GiaoVien gv ON tkg.MaGiaoVien = gv.MaGiaoVien
            WHERE tkg.TenDangNhap = ? AND tkg.MatKhau = ?
        """, (ten_dang_nhap, mat_khau))
        giao_vien = cursor.fetchone()

        if giao_vien:
            user_data = {
                'user_type': 'teacher',
                'ma_giao_vien': giao_vien[0],
                'ho_ten': giao_vien[1],
                'email': giao_vien[2],
                'so_dien_thoai': giao_vien[3],
                'gioi_tinh': giao_vien[4],
                'ngay_sinh': giao_vien[5].strftime('%Y-%m-%d') if giao_vien[5] else '',
                'chuc_vu': giao_vien[6]
            }
            session['user'] = user_data
            return jsonify(
                {'status': 'success', 'message': 'Đăng nhập thành công', 'user': user_data, 'redirect': '/teacher'})

        return jsonify({'status': 'error', 'message': 'Tên đăng nhập hoặc mật khẩu không chính xác'}), 401

    except Exception as e:
        print(f"Login Error: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500
    finally:
        if conn: conn.close()


@app.route('/logout')
def logout():
    session.clear()
    return '''<script>localStorage.removeItem('user'); window.location.href = '/login';</script>'''


# ==========================================
# 2. STUDENT ROUTES & APIs
# ==========================================

@app.route('/dashboard')
def dashboard():
    if 'user' not in session or session['user'].get('user_type') != 'student':
        return redirect(url_for('login_page'))

    user = session['user']
    ma_sinh_vien = user['ma_sinh_vien']

    stats = {'total': 0, 'present': 0, 'absent': 0, 'rate': 0}
    conn = None
    try:
        conn = get_connection()
        if conn:
            cursor = conn.cursor()

            cursor.execute("SELECT COUNT(DISTINCT MaLichHoc) FROM LichHoc_SinhVien WHERE MaSinhVien = ?",
                           (ma_sinh_vien,))
            stats['total'] = cursor.fetchone()[0] or 0

            cursor.execute("SELECT COUNT(*) FROM DiemDanh WHERE MaSinhVien = ? AND TrangThai = N'Có mặt'",
                           (ma_sinh_vien,))
            stats['present'] = cursor.fetchone()[0] or 0

            cursor.execute("SELECT COUNT(*) FROM DiemDanh WHERE MaSinhVien = ? AND TrangThai = N'Vắng mặt'",
                           (ma_sinh_vien,))
            stats['absent'] = cursor.fetchone()[0] or 0

            if stats['total'] > 0:
                stats['rate'] = round((stats['present'] / stats['total']) * 100, 1)
    except Exception as e:
        print(f"Dashboard Error: {e}")
    finally:
        if conn: conn.close()

    return render_template('dashboard.html', user=user, stats=stats)


@app.route('/student/info/<ma_sinh_vien>', methods=['GET'])
def get_student_info(ma_sinh_vien):
    conn = None
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute(
            "SELECT MaSinhVien, HoTen, NgaySinh, GioiTinh, MaLop, MaPhuHuynh, Email, SoDienThoai FROM SinhVien WHERE MaSinhVien = ?",
            (ma_sinh_vien,))
        row = cursor.fetchone()
        if row:
            return jsonify({'status': 'success', 'sinh_vien': {
                'ma_sinh_vien': row[0],
                'ho_ten': row[1],
                'ngay_sinh': row[2].strftime('%Y-%m-%d') if row[2] else '',
                'gioi_tinh': row[3],
                'ma_lop': row[4],
                'email': row[6],
                'so_dien_thoai': row[7]
            }})
        return jsonify({'status': 'error', 'message': 'Không tìm thấy'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})
    finally:
        if conn: conn.close()


@app.route('/student/update_info/<ma_sinh_vien>', methods=['POST'])
def update_student_info(ma_sinh_vien):
    data = request.get_json()
    email = data.get('email')
    sdt = data.get('so_dien_thoai')

    conn = None
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("UPDATE SinhVien SET Email = ?, SoDienThoai = ? WHERE MaSinhVien = ?",
                       (email, sdt, ma_sinh_vien))
        conn.commit()
        return jsonify({'status': 'success', 'message': 'Cập nhật thông tin thành công'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500
    finally:
        if conn: conn.close()


@app.route('/student/schedule/<ma_sinh_vien>', methods=['GET'])
def get_student_schedule(ma_sinh_vien):
    conn = None
    try:
        clean_ma_sv = str(ma_sinh_vien).strip()
        conn = get_connection()
        cursor = conn.cursor()

        query = """
            SELECT DISTINCT
                lh.MaLichHoc, 
                mh.TenMonHoc, 
                lh.NgayHoc, 
                lh.GioBatDau, 
                lh.GioKetThuc, 
                ISNULL(lh.PhongHoc, N'Chưa xếp') as PhongHoc
            FROM LichHoc lh
            JOIN MonHoc mh ON lh.MaMonHoc = mh.MaMonHoc
            JOIN LichDay ld ON lh.MaLichDay = ld.MaLichDay
            JOIN SinhVien sv ON ld.MaLop = sv.MaLop
            WHERE sv.MaSinhVien = ?

            UNION

            SELECT DISTINCT
                lh.MaLichHoc, 
                mh.TenMonHoc, 
                lh.NgayHoc, 
                lh.GioBatDau, 
                lh.GioKetThuc, 
                ISNULL(lh.PhongHoc, N'Chưa xếp') as PhongHoc
            FROM LichHoc lh
            JOIN MonHoc mh ON lh.MaMonHoc = mh.MaMonHoc
            JOIN LichHoc_SinhVien lhs ON lh.MaLichHoc = lhs.MaLichHoc
            WHERE lhs.MaSinhVien = ?

            ORDER BY NgayHoc ASC
        """
        cursor.execute(query, (clean_ma_sv, clean_ma_sv))
        rows = cursor.fetchall()

        schedule = [{
            'ma_lich_hoc': r[0],
            'ten_mon_hoc': r[1],
            'ngay_hoc': r[2].strftime('%d/%m/%Y') if r[2] else '',
            'gio_bat_dau': str(r[3])[:5] if r[3] else '',
            'gio_ket_thuc': str(r[4])[:5] if r[4] else '',
            'phong': r[5]
        } for r in rows]

        return jsonify({'status': 'success', 'schedules': schedule})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500
    finally:
        if conn: conn.close()


@app.route('/student/attendance/<ma_sinh_vien>', methods=['GET'])
def get_student_attendance(ma_sinh_vien):
    conn = None
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT mh.TenMonHoc, dd.ThoiGianDiemDanh, dd.TrangThai, dd.GhiChu
            FROM DiemDanh dd
            JOIN LichHoc lh ON dd.MaLichHoc = lh.MaLichHoc
            JOIN MonHoc mh ON lh.MaMonHoc = mh.MaMonHoc
            WHERE dd.MaSinhVien = ?
            ORDER BY dd.ThoiGianDiemDanh DESC
        """, (ma_sinh_vien,))
        rows = cursor.fetchall()

        data = [{
            'ten_mon_hoc': r[0],
            'thoi_gian': r[1].strftime('%d/%m/%Y %H:%M') if r[1] else 'Chưa có',
            'trang_thai': r[2],
            'ghi_chu': r[3] if r[3] else ''
        } for r in rows]
        return jsonify({'status': 'success', 'data': data})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})
    finally:
        if conn: conn.close()


# ==========================================
# 3. TEACHER ROUTES & APIs
# ==========================================

@app.route('/teacher')
def teacher_dashboard():
    if 'user' not in session or session['user'].get('user_type') != 'teacher':
        return redirect(url_for('login_page'))
    return render_template('teacher_dashboard.html', user=session['user'])


@app.route('/teacher/info/<ma_giao_vien>', methods=['GET'])
def get_teacher_info(ma_giao_vien):
    conn = None
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute(
            "SELECT MaGiaoVien, HoTen, NgaySinh, GioiTinh, Email, SoDienThoai, Chucvu FROM GiaoVien WHERE MaGiaoVien = ?",
            (ma_giao_vien,))
        row = cursor.fetchone()
        if row:
            return jsonify({'status': 'success', 'giao_vien': {
                'ho_ten': row[1],
                'email': row[4],
                'so_dien_thoai': row[5],
                'chuc_vu': row[6]
            }})
        return jsonify({'status': 'error', 'message': 'Not found'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})
    finally:
        if conn: conn.close()


@app.route('/teacher/update_info/<ma_giao_vien>', methods=['POST'])
def update_teacher_info(ma_giao_vien):
    data = request.get_json()
    email = data.get('email')
    sdt = data.get('so_dien_thoai')
    conn = None
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("UPDATE GiaoVien SET Email = ?, SoDienThoai = ? WHERE MaGiaoVien = ?",
                       (email, sdt, ma_giao_vien))
        conn.commit()
        return jsonify({'status': 'success', 'message': 'Cập nhật thông tin thành công'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500
    finally:
        if conn: conn.close()


@app.route('/api/teacher/update-attendance-note', methods=['POST'])
def update_attendance_note():
    data = request.get_json()
    ma_sv = data.get('ma_sinh_vien')
    ma_lh = data.get('ma_lich_hoc')
    ghi_chu = data.get('ghi_chu')

    conn = None
    try:
        conn = get_connection()
        cursor = conn.cursor()

        cursor.execute("SELECT 1 FROM DiemDanh WHERE MaSinhVien = ? AND MaLichHoc = ?", (ma_sv, ma_lh))
        if not cursor.fetchone():
            cursor.execute("""
                INSERT INTO DiemDanh (MaSinhVien, MaLichHoc, ThoiGianDiemDanh, TrangThai, GhiChu)
                VALUES (?, ?, GETDATE(), N'Vắng mặt', ?)
            """, (ma_sv, ma_lh, ghi_chu))
        else:
            cursor.execute("UPDATE DiemDanh SET GhiChu = ? WHERE MaSinhVien = ? AND MaLichHoc = ?",
                           (ghi_chu, ma_sv, ma_lh))

        conn.commit()
        return jsonify({'status': 'success', 'message': 'Lưu ghi chú thành công!'})
    except Exception as e:
        print(f"Lỗi update note: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500
    finally:
        if conn: conn.close()


@app.route('/api/teacher/dashboard/<ma_giao_vien>', methods=['GET'])
def get_teacher_stats(ma_giao_vien):
    conn = None
    try:
        conn = get_connection()
        cursor = conn.cursor()

        cursor.execute("SELECT COUNT(DISTINCT MaLop) FROM LichDay WHERE MaGiaoVien = ?", (ma_giao_vien,))
        total_classes = cursor.fetchone()[0] or 0

        cursor.execute("SELECT COUNT(*) FROM LichDay WHERE MaGiaoVien = ?", (ma_giao_vien,))
        total_sessions = cursor.fetchone()[0] or 0

        cursor.execute("""
            SELECT COUNT(*) FROM DiemDanh dd
            JOIN LichHoc lh ON dd.MaLichHoc = lh.MaLichHoc
            JOIN LichDay ld ON lh.MaLichDay = ld.MaLichDay
            WHERE ld.MaGiaoVien = ? AND dd.TrangThai = N'Vắng mặt'
        """, (ma_giao_vien,))
        absent_records = cursor.fetchone()[0] or 0

        return jsonify({'status': 'success', 'stats': {
            'total_classes': total_classes,
            'total_sessions': total_sessions,
            'absent_records': absent_records
        }})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})
    finally:
        if conn: conn.close()


@app.route('/teacher/classes/<ma_giao_vien>', methods=['GET'])
def get_teacher_schedule_list(ma_giao_vien):
    conn = None
    try:
        conn = get_connection()
        cursor = conn.cursor()

        query = """
            SELECT 
                ISNULL(mh.TenMonHoc, N'Môn chưa định nghĩa') as TenMonHoc, 
                ISNULL(l.TenLop, N'Lớp chưa định nghĩa') as TenLop, 
                ld.NgayDay, 
                ld.GioBatDau, 
                ld.GioKetThuc
            FROM LichDay ld
            LEFT JOIN MonHoc mh ON ld.MaMonHoc = mh.MaMonHoc
            LEFT JOIN Lop l ON ld.MaLop = l.MaLop
            WHERE ld.MaGiaoVien = ?
            ORDER BY ld.NgayDay DESC
        """
        cursor.execute(query, (ma_giao_vien,))
        rows = cursor.fetchall()

        classes = [{
            'ten_mon_hoc': r[0],
            'ten_lop': r[1],
            'ngay_day': r[2].strftime('%d/%m/%Y') if r[2] else '',
            'gio_bat_dau': str(r[3])[:5] if r[3] else '',
            'gio_ket_thuc': str(r[4])[:5] if r[4] else ''
        } for r in rows]
        return jsonify({'status': 'success', 'classes': classes})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})
    finally:
        if conn: conn.close()


@app.route('/api/teacher/absences/<ma_giao_vien>', methods=['GET'])
def get_teacher_absences(ma_giao_vien):
    conn = None
    try:
        conn = get_connection()
        cursor = conn.cursor()
        query = """
            SELECT TOP 10 sv.MaSinhVien, sv.HoTen, mh.TenMonHoc, lh.NgayHoc, dd.GhiChu
            FROM DiemDanh dd
            JOIN SinhVien sv ON dd.MaSinhVien = sv.MaSinhVien
            JOIN LichHoc lh ON dd.MaLichHoc = lh.MaLichHoc
            JOIN LichDay ld ON lh.MaLichDay = ld.MaLichDay
            JOIN MonHoc mh ON ld.MaMonHoc = mh.MaMonHoc
            WHERE ld.MaGiaoVien = ? AND dd.TrangThai = N'Vắng mặt'
            ORDER BY lh.NgayHoc DESC
        """
        cursor.execute(query, (ma_giao_vien,))
        rows = cursor.fetchall()

        absences = [{
            'ma_sinh_vien': r[0],
            'ho_ten': r[1],
            'ten_mon_hoc': r[2],
            'ngay_hoc': r[3].strftime('%d/%m/%Y') if r[3] else '',
            'ghi_chu': r[4] if r[4] else 'Không phép'
        } for r in rows]
        return jsonify({'status': 'success', 'absences': absences})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})
    finally:
        if conn: conn.close()


@app.route('/api/teacher/classes-for-report/<ma_giao_vien>', methods=['GET'])
def get_classes_for_report(ma_giao_vien):
    conn = None
    try:
        conn = get_connection()
        cursor = conn.cursor()
        query = """
            SELECT lh.MaLichHoc, lh.NgayHoc, mh.TenMonHoc, l.TenLop
            FROM LichHoc lh
            JOIN LichDay ld ON lh.MaLichDay = ld.MaLichDay
            JOIN MonHoc mh ON ld.MaMonHoc = mh.MaMonHoc
            JOIN Lop l ON ld.MaLop = l.MaLop
            WHERE ld.MaGiaoVien = ?
            ORDER BY lh.NgayHoc DESC
        """
        cursor.execute(query, (ma_giao_vien,))
        rows = cursor.fetchall()

        classes = [{
            'ma_lich_hoc': r[0],
            'ngay_hoc': r[1].strftime('%d/%m/%Y') if r[1] else '',
            'ten_mon_hoc': r[2],
            'ten_lop': r[3]
        } for r in rows]
        return jsonify({'status': 'success', 'classes': classes})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})
    finally:
        if conn: conn.close()


@app.route('/api/teacher/attendance-report/<ma_lich_hoc>', methods=['GET'])
def get_attendance_report_detail(ma_lich_hoc):
    conn = None
    try:
        conn = get_connection()
        cursor = conn.cursor()

        # 1. LẤY DỮ LIỆU ĐỂ HIỂN THỊ WEB (Sử dụng UNION để lấy cả 2 nguồn)
        query = """
            SELECT sv.MaSinhVien, sv.HoTen, sv.MaLop, 
                   ISNULL(dd.TrangThai, N'Chưa điểm danh') as TrangThai, 
                   dd.ThoiGianDiemDanh, dd.GhiChu
            FROM SinhVien sv
            JOIN LichDay ld ON sv.MaLop = ld.MaLop
            JOIN LichHoc lh ON ld.MaLichDay = lh.MaLichDay
            LEFT JOIN DiemDanh dd ON sv.MaSinhVien = dd.MaSinhVien AND dd.MaLichHoc = lh.MaLichHoc
            WHERE lh.MaLichHoc = ?

            UNION

            SELECT sv.MaSinhVien, sv.HoTen, sv.MaLop,
                   ISNULL(dd.TrangThai, N'Chưa điểm danh'),
                   dd.ThoiGianDiemDanh, dd.GhiChu
            FROM LichHoc_SinhVien lhs
            JOIN SinhVien sv ON lhs.MaSinhVien = sv.MaSinhVien
            LEFT JOIN DiemDanh dd ON sv.MaSinhVien = dd.MaSinhVien AND dd.MaLichHoc = lhs.MaLichHoc
            WHERE lhs.MaLichHoc = ?
        """
        cursor.execute(query, (ma_lich_hoc, ma_lich_hoc))
        rows = cursor.fetchall()

        students = [{
            'ma_sinh_vien': r[0],
            'ho_ten': r[1],
            'ma_lop': r[2],
            'trang_thai': r[3],
            'thoi_gian_diem_danh': r[4].strftime('%H:%M:%S') if r[4] else '',
            'ghi_chu': r[5] if r[5] else ''
        } for r in rows]

        return jsonify({'status': 'success', 'students': students})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})
    finally:
        if conn: conn.close()


# --- SỬA LỖI EXCEL: ÁP DỤNG CÂU QUERY "UNION" GIỐNG HỆT WEB ---
@app.route('/api/teacher/export-attendance/<ma_lich_hoc>', methods=['GET'])
def export_attendance_excel(ma_lich_hoc):
    if not OPENPYXL_AVAILABLE:
        return "Server chưa cài thư viện openpyxl. Vui lòng chạy: pip install openpyxl", 500

    conn = None
    try:
        conn = get_connection()
        cursor = conn.cursor()

        # Lấy thông tin Header
        cursor.execute("""
            SELECT mh.TenMonHoc, l.TenLop, lh.NgayHoc 
            FROM LichHoc lh 
            JOIN LichDay ld ON lh.MaLichDay = ld.MaLichDay 
            JOIN MonHoc mh ON ld.MaMonHoc = mh.MaMonHoc 
            JOIN Lop l ON ld.MaLop = l.MaLop
            WHERE lh.MaLichHoc = ?
        """, (ma_lich_hoc,))
        info = cursor.fetchone()

        if not info:
            return "Không tìm thấy thông tin buổi học", 404

        ten_mon, ten_lop, ngay_hoc = info[0], info[1], info[2].strftime('%d/%m/%Y')

        # === FIX LỖI TẠI ĐÂY ===
        # Sử dụng UNION để lấy đủ cả 2 loại sinh viên (như trên Web)
        # 1. Sinh viên theo Lớp
        # 2. Sinh viên đăng ký riêng (LichHoc_SinhVien)
        query_data = """
            SELECT sv.MaSinhVien, sv.HoTen, sv.MaLop, ISNULL(dd.TrangThai, N'Chưa điểm danh'), ISNULL(dd.GhiChu, '')
            FROM SinhVien sv 
            JOIN LichDay ld ON sv.MaLop = ld.MaLop 
            JOIN LichHoc lh ON ld.MaLichDay = lh.MaLichDay
            LEFT JOIN DiemDanh dd ON sv.MaSinhVien = dd.MaSinhVien AND dd.MaLichHoc = lh.MaLichHoc
            WHERE lh.MaLichHoc = ?

            UNION

            SELECT sv.MaSinhVien, sv.HoTen, sv.MaLop, ISNULL(dd.TrangThai, N'Chưa điểm danh'), ISNULL(dd.GhiChu, '')
            FROM LichHoc_SinhVien lhs
            JOIN SinhVien sv ON lhs.MaSinhVien = sv.MaSinhVien
            LEFT JOIN DiemDanh dd ON sv.MaSinhVien = dd.MaSinhVien AND dd.MaLichHoc = lhs.MaLichHoc
            WHERE lhs.MaLichHoc = ?
        """
        # Lưu ý: Cần truyền tham số 2 lần cho 2 dấu ?
        cursor.execute(query_data, (ma_lich_hoc, ma_lich_hoc))
        students = cursor.fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = "Diem Danh"

        ws['A1'] = "BÁO CÁO ĐIỂM DANH"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:F1')
        ws['A1'].alignment = Alignment(horizontal='center')

        ws['A2'] = f"Môn học: {ten_mon} - Lớp: {ten_lop}"
        ws.merge_cells('A2:F2')
        ws['A2'].alignment = Alignment(horizontal='center')

        ws['A3'] = f"Ngày học: {ngay_hoc}"
        ws.merge_cells('A3:F3')
        ws['A3'].alignment = Alignment(horizontal='center')

        headers = ['STT', 'Mã SV', 'Họ và Tên', 'Lớp', 'Trạng Thái', 'Ghi Chú']
        ws.append([])
        ws.append(headers)

        for cell in ws[5]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

        for idx, sv in enumerate(students, 1):
            ws.append([idx, sv[0], sv[1], sv[2], sv[3], sv[4]])
            status_cell = ws.cell(row=5 + idx, column=5)
            if sv[3] == 'Có mặt':
                status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif sv[3] == 'Vắng mặt':
                status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 30

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"DiemDanh_{ten_lop}_{ngay_hoc.replace('/', '')}.xlsx"
        return send_file(output, download_name=filename, as_attachment=True,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        print(f"Excel Error: {e}")
        return str(e), 500
    finally:
        if conn: conn.close()


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)